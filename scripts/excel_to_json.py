"""
Excel → JSON converter for:
  FY_CUT_OFF__Placement_data__Fee__Hostel_2025_-_26.xlsx

Sheets handled:
  FY 27     → cut-off data, bus routes, hostel fees, tuition fees
  TPO Data  → last 3-year placement summary
  Sheet2    → company visited lists (2022-23, 2023-24, 2024-25)
  Sheet4    → placement data per academic year (separate layout)
  Doc       → documents required for FY and DSY admissions
  Sheet1    → programs offered (B.Tech / BBA / BCA / M.Tech)

Usage:
  pip install openpyxl
  python excel_to_json.py
  # Reads from same directory; writes output.json next to the script.
"""

import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

# ── helpers ────────────────────────────────────────────────────────────────────

def clean(v):
    """Normalise a cell value: strip strings, convert '-' / None → None."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v in ("-", "--", ""):
            return None
    return v


def row_vals(row):
    return [clean(c) for c in row]


# ── Sheet: Sheet1 – Programs offered ──────────────────────────────────────────

def parse_sheet1(ws):
    programs = {"btech": [], "ug_other": [], "mtech": []}
    current_section = None

    for row in ws.iter_rows(values_only=True):
        r = row_vals(row)
        first = r[0]

        if first is None:
            continue
        if isinstance(first, str) and "B.Tech" in first:
            current_section = "btech"
            continue
        if isinstance(first, str) and "BBA" in first:
            current_section = "ug_other"
            continue
        if isinstance(first, str) and "M.Tech" in first:
            current_section = "mtech"
            continue

        # skip pure header rows
        if first in ("Sr.No.", "Sr. No"):
            continue

        if isinstance(first, int) and current_section:
            if current_section == "btech":
                programs["btech"].append({
                    "sr_no": r[0],
                    "program": r[1],
                    "program_code": r[2],
                    "sanctioned_intake": r[3],
                    "year_of_starting": r[4],
                })
            elif current_section == "ug_other":
                programs["ug_other"].append({
                    "sr_no": r[0],
                    "program": r[1],
                    "program_code": r[2],
                    "sanctioned_intake": r[3],
                    "year_of_starting": r[4],
                })
            elif current_section == "mtech":
                programs["mtech"].append({
                    "sr_no": r[0],
                    "program": r[1],
                    "sanctioned_intake": r[2],
                    "year_of_starting": r[3],
                })

    return programs


# ── Sheet: Doc – Required documents ───────────────────────────────────────────

def parse_doc(ws):
    categories = ["OPEN / EBC / TFWS / EWS", "SC / ST",
                  "VJ / VJNT / NT1 / NT2 / NT3", "OBC / SBC / SEBC"]

    sections = {}
    current_section = None
    current_docs = []

    for row in ws.iter_rows(values_only=True):
        r = row_vals(row)
        first = r[0]

        if first is None:
            continue

        # detect section title rows
        if isinstance(first, str) and "List of Documents" in first:
            if current_section and current_docs:
                sections[current_section] = current_docs
            current_section = first
            current_docs = []
            continue

        # skip pure header rows
        if first == "Documents":
            continue
        if first is None and r[1] in categories:
            continue

        if current_section and first is not None:
            doc_entry = {"document": first}
            for i, cat in enumerate(categories):
                doc_entry[cat] = r[i + 1] if i + 1 < len(r) else None
            current_docs.append(doc_entry)

    if current_section and current_docs:
        sections[current_section] = current_docs

    return sections


# ── Sheet: TPO Data – 3-year placement summary ────────────────────────────────

def parse_tpo_data(ws):
    rows = [row_vals(r) for r in ws.iter_rows(values_only=True)
            if any(c is not None for c in r)]

    # rows[3] = year headers, rows[4] = sub-headers, rows[5..] = branch data
    years = []
    for v in rows[3]:
        if isinstance(v, str) and "Acadamic Year" in v:
            years.append(v.replace("Acadamic Year:- ", "").strip())

    branches = []
    total_row = {}
    company_visits = {}

    for r in rows[5:]:
        branch = r[0]
        if branch is None:
            continue
        if "Overall Total" in str(branch):
            # placed counts sit in cols 3, 7, 11
            total_row = {
                years[0] if len(years) > 0 else "2025-26": {"students_placed_offers": r[3]},
                years[1] if len(years) > 1 else "2024-25": {"total_students": r[6], "students_placed_offers": r[7]},
                years[2] if len(years) > 2 else "2023-24": {"total_students": r[10], "students_placed_offers": r[11]},
            }
            continue
        if "Company Visited" in str(branch):
            company_visits = {
                years[0] if len(years) > 0 else "2025-26": r[2],
                years[1] if len(years) > 1 else "2024-25": r[6],
                years[2] if len(years) > 2 else "2023-24": r[10],
            }
            continue

        entry = {"branch": branch.strip()}
        for i, yr in enumerate(years):
            base = 2 + i * 4
            entry[yr] = {
                "total_students": r[base],
                "students_placed_offers": r[base + 1],
                "avg_salary_lpa": r[base + 2],
                "highest_salary_lpa": r[base + 3],
            }
        branches.append(entry)

    return {
        "placement_summary_by_branch": branches,
        "overall_totals": total_row,
        "companies_visited": company_visits,
    }


# ── Sheet: Sheet4 – placement data alternate layout ───────────────────────────

def parse_sheet4(ws):
    rows = [row_vals(r) for r in ws.iter_rows(values_only=True)
            if any(c is not None for c in r)]

    result = {}
    current_year = None
    current_data = []

    for r in rows:
        first = r[0]
        # detect year heading
        yr_match = None
        for cell in r:
            if isinstance(cell, str) and "Acadamic Year" in cell:
                yr_match = cell.replace("Acadamic Year:- ", "").strip()
                break

        if yr_match:
            if current_year and current_data:
                result[current_year] = current_data
            current_year = yr_match
            current_data = []
            continue

        if first == "Branch" or first is None:
            continue
        if "Overall Total" in str(first):
            if current_year is not None:
                total = {"total_students": r[1], "students_placed_offers": r[2]}
                result.setdefault(current_year + "_summary", {})["overall_total"] = total
            continue
        if "Company Visited" in str(first):
            if current_year is not None:
                result.setdefault(current_year + "_summary", {})["companies_visited"] = r[1]
            continue

        if current_year and isinstance(first, str):
            current_data.append({
                "branch": first.strip(),
                "total_students": r[1],
                "students_placed_offers": r[2],
                "avg_salary_lpa": r[3],
                "highest_salary_lpa": r[4],
            })

    if current_year and current_data:
        result[current_year] = current_data

    return result


# ── Sheet: Sheet2 – company visited lists ─────────────────────────────────────

def parse_sheet2(ws):
    rows = [row_vals(r) for r in ws.iter_rows(values_only=True)
            if any(c is not None for c in r)]

    result = {}
    current_year = None
    current_companies = []

    for r in rows:
        first = r[0]
        if isinstance(first, str) and "Company Visited List" in first:
            if current_year and current_companies:
                result[current_year] = current_companies
            current_year = first.replace("Company Visited List ", "").strip()
            current_companies = []
            continue

        # skip header rows and institute name rows
        if first == "Company Name" or (isinstance(first, str) and "Sanstha" in first):
            continue
        if isinstance(first, str) and ("Annasaheb" in first or "Autonomous" in first
                                        or "Changing Lives" in first):
            continue

        if current_year and first is not None:
            # each row can have up to 3 companies (cols 0-2, 4-6, 8-10)
            for start in [0, 4, 8]:
                name = r[start] if start < len(r) else None
                ind = r[start + 1] if start + 1 < len(r) else None
                branches = r[start + 2] if start + 2 < len(r) else None
                if name and name != "Company Name":
                    current_companies.append({
                        "company_name": name,
                        "industry_vertical": ind,
                        "eligible_branches": branches,
                    })

    if current_year and current_companies:
        result[current_year] = current_companies

    return result


# ── Sheet: FY 27 – cut-off / bus / hostel / fees ──────────────────────────────

CUTOFF_CATEGORY_SETS = [
    ["VJ", "NT-1", "NT-2", "NT-3"],
    ["OBC", "SEBC", "OPEN", "SC"],
    ["ST", "DEF", "TFWS", "EWS"],
]

def build_empty_course():
    return {cat: {"G": {"merit_no": None, "merit_marks": None},
                  "L": {"merit_no": None, "merit_marks": None}}
            for group in CUTOFF_CATEGORY_SETS for cat in group}


def parse_cutoff_block(rows, col_offset):
    """Parse one side (left or right) of a course block from the raw rows list.
    col_offset = 0 for left half, 13 for right half."""
    course_data = build_empty_course()

    # rows[0]: sr_no at col0+col_offset, course at col1+col_offset
    # rows[1]: Merit No / Merit Marks sub-headers (skip)
    # rows[2..9]: actual data

    # Map category column positions to category names
    # For each category-set row there are 4 categories, each with 2 cols
    cat_col_map = {
        0: "VJ", 2: "NT-1", 4: "NT-2", 6: "NT-3",   # relative within the 8-col category span
    }

    i = 0
    gender = None
    current_cats = None

    while i < len(rows):
        r = rows[i]
        base = col_offset + 3  # VJ merit_no starts here relative to sheet

        cell2 = clean(r[col_offset + 2]) if col_offset + 2 < len(r) else None

        if cell2 == "G" or cell2 == "L":
            gender = cell2
            # read 4 categories from this row
            cats = current_cats or CUTOFF_CATEGORY_SETS[0]
            for j, cat in enumerate(cats):
                mn_col = base + j * 2
                mm_col = mn_col + 1
                mn = clean(r[mn_col]) if mn_col < len(r) else None
                mm = clean(r[mm_col]) if mm_col < len(r) else None
                if cat in course_data:
                    course_data[cat][gender]["merit_no"] = mn
                    course_data[cat][gender]["merit_marks"] = mm

        elif cell2 == "Category":
            # next four non-None values after col3 are the category names
            cats = []
            for j in range(4):
                c = clean(r[base + j * 2]) if base + j * 2 < len(r) else None
                if c:
                    cats.append(c)
            if cats:
                current_cats = cats

        i += 1

    return course_data


def parse_fy27(ws):
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(row_vals(row))

    # ── identify section boundaries ───────────────────────────────────────────
    bus_start = None
    hostel_start = None
    for idx, r in enumerate(all_rows):
        if r[0] == "Bus Facility":
            bus_start = idx
        if r[0] == "Hostel":
            hostel_start = idx

    cutoff_rows = all_rows[:bus_start] if bus_start else all_rows
    bus_rows = all_rows[bus_start:hostel_start] if bus_start and hostel_start else []
    hostel_rows = all_rows[hostel_start:] if hostel_start else []

    # ── 1. Cut-off data ───────────────────────────────────────────────────────
    courses = {}
    current_block_rows = []
    left_meta = None   # (sr_no, course_name)
    right_meta = None

    def flush_block(left_meta, right_meta, block_rows):
        """Parse one double-wide block (9 data rows) and store into `courses`."""
        if left_meta:
            courses[left_meta] = {
                "sr_no": left_meta[0],
                "course": left_meta[1],
                "cutoff": parse_cutoff_block(block_rows, col_offset=0),
            }
        if right_meta:
            courses[right_meta] = {
                "sr_no": right_meta[0],
                "course": right_meta[1],
                "cutoff": parse_cutoff_block(block_rows, col_offset=12),
            }

    # Sheet layout note:
    #   A row with an int in col 0 marks the start of a new course block.
    #   For the FIRST block (Mech) this row ALSO carries the first 'G'
    #   data (col2='G', data in cols 3-10).
    #   For later blocks (CSE onward) the start row is a pure header
    #   (col2=None), followed by a Merit-No sub-header row, then data.
    #   We accumulate only rows where col2 is 'G', 'L', or 'Category',
    #   so header/sub-header rows are naturally skipped.

    for r in cutoff_rows:
        left_sr = clean(r[0])
        right_sr = clean(r[12]) if len(r) > 12 else None
        col2    = clean(r[2])   if len(r) > 2  else None

        if isinstance(left_sr, int):
            if current_block_rows:
                flush_block(left_meta, right_meta, current_block_rows)
            left_meta  = (left_sr, clean(r[1]))
            right_meta = (right_sr, clean(r[13])) if right_sr is not None else None
            current_block_rows = []
            # Fall through: if col2 is 'G'/'L'/'Category' this very row
            # carries data (happens for the Mech block)

        if col2 in ("G", "L", "Category"):
            current_block_rows.append(r)

    if current_block_rows:
        flush_block(left_meta, right_meta, current_block_rows)

    # ── 2. Bus facility ───────────────────────────────────────────────────────
    routes = {}

    def parse_route_table(rows, base_col, route_name_row_idx, rows_list):
        """Extract stops + fees for one route table."""
        route_name = clean(rows_list[route_name_row_idx][base_col])
        if not route_name:
            return None, []
        stops = []
        for r in rows_list[route_name_row_idx + 1:]:
            sr = clean(r[base_col]) if base_col < len(r) else None
            stop = clean(r[base_col + 1]) if base_col + 1 < len(r) else None
            fee = clean(r[base_col + 4]) if base_col + 4 < len(r) else None
            if sr is None or stop is None:
                break
            stops.append({"sr_no": sr, "stop": stop, "fee": fee})
        return route_name, stops

    # Bus rows layout: header row per route block, then stop rows
    # Col offsets for route tables: 0, 5, 11, 17 (approx based on data)
    # Actually from data: 4 route tables per row block, col offsets 0,5,11,17
    if bus_rows:
        # Find the header rows (those containing route names like "Ashta-Palus")
        route_header_rows = [i for i, r in enumerate(bus_rows)
                             if isinstance(clean(r[1]), str) and "Ashta" in str(clean(r[1]) or "")
                             or isinstance(clean(r[0]), str) and "Sr.No" == clean(r[0])]

        # Parse based on actual structure: route name is in col1, col5, col11, col17 of header rows
        # More reliable: scan for "Sr.No" rows which mark route headers
        header_indices = [i for i, r in enumerate(bus_rows)
                          if clean(r[0]) == "Sr.No"]

        for hi in header_indices:
            r = bus_rows[hi]
            # Parse up to 4 route tables from this header row
            # Col 0:"Sr.No", col1: route_name, col4: Fee -> table 1
            # Col 5:"Sr.No", col6: route_name, col9: Fee -> table 2  (skip col 10=None)
            # Col 11:"Sr.No", col12: route_name, col15: Fee -> table 3
            # Col 16:"Sr.No", col17: route_name, col20: Fee -> table 4
            table_configs = [(0, 1, 4), (5, 6, 9), (11, 12, 15), (16, 17, 20)]
            for sr_col, name_col, fee_col in table_configs:
                route_name = clean(r[name_col]) if name_col < len(r) else None
                if not route_name or route_name == "Sr.No":
                    continue
                stops = []
                # Collect stop rows until we hit another header or run out
                for rr in bus_rows[hi + 1:]:
                    sr = clean(rr[sr_col]) if sr_col < len(rr) else None
                    stop_name = clean(rr[name_col]) if name_col < len(rr) else None
                    fee = clean(rr[fee_col]) if fee_col < len(rr) else None
                    if clean(rr[0]) == "Sr.No":
                        break
                    if sr is None:
                        continue
                    if stop_name:
                        stops.append({"sr_no": sr, "stop": stop_name, "fee": fee})
                if stops:
                    routes[f"Ashta-{route_name}" if not route_name.startswith("Ashta") else route_name] = stops

    # ── 3. Hostel & Fees ──────────────────────────────────────────────────────
    hostels = []
    qualifying_criteria = []
    fees = []

    if hostel_rows:
        # Row layout (confirmed from raw data inspection):
        # row 0: section title header ("Hostel" | "12th Group..." | "Fees For A.Y. 2025-26")
        # row 1: sub-headers ("New ladies hostel" | criteria text | "Category"/"FY"/"DSE" headers)
        # rows 2-5: data rows
        #
        # Column indices (0-based):
        #   col 0  = hostel name
        #   col 4  = hostel annual fee
        #   col 6  = qualifying criteria description
        #   col 13 = qualifying criteria marks (e.g. "134/135")
        #   col 15 = fee category name
        #   col 19 = FY fee amount
        #   col 21 = DSE fee amount
        # row 1 contains the "Category / FY / DSE" sub-header — skip it for fee data

        for r in hostel_rows[1:]:   # skip only row 0 (section title "Hostel")
            hostel_name = clean(r[0]) if len(r) > 0 else None
            hostel_fee  = clean(r[4]) if len(r) > 4 else None
            criteria_desc  = clean(r[6]) if len(r) > 6 else None
            criteria_marks = clean(r[13]) if len(r) > 13 else None
            fee_cat = clean(r[15]) if len(r) > 15 else None
            fee_fy  = clean(r[19]) if len(r) > 19 else None
            fee_dse = clean(r[21]) if len(r) > 21 else None

            if hostel_name:
                hostels.append({"hostel": hostel_name, "annual_fee": hostel_fee})
            if criteria_desc:
                qualifying_criteria.append({
                    "description": criteria_desc,
                    "marks_out_of": criteria_marks,
                })
            if fee_cat and fee_cat not in ("Category", "Category "):
                fees.append({
                    "category": fee_cat.strip(),
                    "fy_fee": fee_fy,
                    "dse_fee": fee_dse,
                })

    return {
        "cutoff_data": list(courses.values()),
        "bus_routes": routes,
        "hostels": hostels,
        "qualifying_criteria_12th": qualifying_criteria,
        "tuition_fees_2025_26": fees,
    }


# ── Main ───────────────────────────────────────────────────────────────────────

def main(input_path: str, output_path: str):
    wb = load_workbook(input_path, read_only=True, data_only=True)

    result = {
        "source_file": Path(input_path).name,
        "college": "Annasaheb Dange College of Engineering and Technology, Ashta",
        "academic_year": "2025-26",
        "sheets": {},
    }

    sheet_parsers = {
        "FY 27":    ("fy27_cutoff_bus_hostel_fees", parse_fy27),
        "TPO Data ": ("placement_3yr_summary",       parse_tpo_data),
        "Sheet2":   ("company_visited_lists",         parse_sheet2),
        "Sheet4":   ("placement_by_year_detail",      parse_sheet4),
        "Doc":      ("admission_documents",            parse_doc),
        "Sheet1":   ("programs_offered",              parse_sheet1),
    }

    for sheet_name, (key, parser) in sheet_parsers.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            try:
                result["sheets"][key] = parser(ws)
                print(f"  ✓  {sheet_name:20s} → {key}")
            except Exception as exc:
                result["sheets"][key] = {"error": str(exc)}
                print(f"  ✗  {sheet_name:20s} → ERROR: {exc}", file=sys.stderr)
        else:
            print(f"  –  {sheet_name:20s} not found, skipping")

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"\nJSON written to: {output_path}")
    print(f"Total size: {Path(output_path).stat().st_size / 1024:.1f} KB")


if __name__ == "__main__":
    default_input = "FY_CUT_OFF__Placement_data__Fee__Hostel_2025_-_26.xlsx"
    default_output = "output.json"

    inp = sys.argv[1] if len(sys.argv) > 1 else default_input
    out = sys.argv[2] if len(sys.argv) > 2 else default_output

    print(f"Parsing: {inp}")
    main(inp, out)
